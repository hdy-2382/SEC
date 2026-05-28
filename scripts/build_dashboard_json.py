"""
build_dashboard_json.py
-----------------------
업체가 보낸 .xlsx 파일을 읽어서 data/dashboard.json 으로 변환한다.

규칙:
  - data/raw/ 폴더 내에서 가장 최근에 수정된 .xlsx 파일을 입력으로 사용
  - 엑셀 시트:
      "일일평가"  → daily 배열
      "에러로그"  → errors 배열
    (시트명이 정확히 일치하지 않으면 부분 매칭으로 찾는다)
  - 출력: data/dashboard.json (UTF-8, 한글 그대로)

읽기 방식:
  - 1차: openpyxl (빠름, 일반 xlsx 전용)
  - 2차 폴백: xlwings (Excel 실행, DRM 보호된 xlsx도 처리 가능)
    → Windows + Excel 설치 필요. openpyxl이 zipfile.BadZipFile로 실패하면 자동 전환.

기대 컬럼:
  일일평가 시트  : 평가일, 입실인원, 주평가내용, 일일평가, 일일에러, 연속성공, 비고
  에러로그 시트  : No, 발생일, 시각, 회차, 코드, 유형, 상세, 원인, 조치, 결과, 담당
"""

from __future__ import annotations

import json
import re
import zipfile
from datetime import date, datetime, time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUT_PATH = ROOT / "data" / "dashboard.json"


# ── 컬럼 헤더 매핑 ─────────────────────────────────────────────
# 한글 헤더에서 공백/괄호 제거 후 매칭한다 (오타 방지)
DAILY_FIELD_ALIASES = {
    "date":      ["평가일", "일자", "date"],
    "personnel": ["입실인원", "입실자", "인원", "personnel"],
    "activity":  ["주평가내용", "평가내용", "내용", "activity"],
    "total":     ["일일평가", "평가횟수", "사이클", "total"],
    "errors":    ["일일에러", "에러", "errors"],
    "streak":    ["연속성공", "연속", "streak"],
    "notes":     ["비고", "메모", "notes"],
}
ERROR_FIELD_ALIASES = {
    "no":     ["no", "번호", "순번", "no."],
    "date":   ["발생일", "일자", "date"],
    "time":   ["시각", "시간", "time"],
    "cycle":  ["회차", "사이클", "cycle"],
    "code":   ["코드", "code"],
    "type":   ["유형", "타입", "type"],
    "detail": ["상세", "상세내용", "detail"],
    "cause":  ["원인", "cause"],
    "action": ["조치", "조치사항", "action"],
    "result": ["결과", "조치결과", "result"],
    "owner":  ["담당", "담당자", "owner"],
}

DAILY_SHEET_KEYWORDS  = ["일일평가", "일일", "daily"]
ERRORS_SHEET_KEYWORDS = ["에러로그", "에러", "error"]


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).lower()


def _find_name(names: list[str], keywords: list[str]) -> str | None:
    """시트명 리스트에서 keywords 중 하나라도 포함된 첫 이름을 반환."""
    norm_map = {n: _norm(n) for n in names}
    for kw in keywords:
        kw_n = _norm(kw)
        for name in names:
            if kw_n in norm_map[name]:
                return name
    return None


def _build_column_map(header_row, aliases: dict[str, list[str]]) -> dict[str, int]:
    """헤더 행을 보고 {필드명: 컬럼인덱스} 매핑을 만든다."""
    norm_cells = [_norm(c) for c in header_row]
    col_map: dict[str, int] = {}
    for field, candidates in aliases.items():
        for cand in candidates:
            cand_n = _norm(cand)
            for idx, cell in enumerate(norm_cells):
                if cell == cand_n or (cand_n and cand_n in cell):
                    col_map[field] = idx
                    break
            if field in col_map:
                break
    return col_map


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    return str(v).strip()


def _cell_to_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _parse_daily(rows: list[list]) -> list[dict]:
    if not rows:
        return []
    header, *body = rows
    cmap = _build_column_map(header, DAILY_FIELD_ALIASES)
    if "date" not in cmap or "total" not in cmap:
        raise SystemExit(
            f"[일일평가] 시트에서 필수 컬럼(평가일/일일평가)을 찾지 못했습니다. "
            f"감지된 헤더: {[str(h) for h in header]}"
        )

    out = []
    for row in body:
        if row is None or all(c is None or c == "" for c in row):
            continue
        date_val = _cell_to_str(row[cmap["date"]])
        if not date_val:
            continue
        out.append({
            "date":      date_val,
            "personnel": _cell_to_str(row[cmap["personnel"]]) if "personnel" in cmap else "",
            "activity":  _cell_to_str(row[cmap["activity"]])  if "activity"  in cmap else "",
            "total":     _cell_to_int(row[cmap["total"]]),
            "errors":    _cell_to_int(row[cmap["errors"]])    if "errors"    in cmap else 0,
            "streak":    _cell_to_int(row[cmap["streak"]])    if "streak"    in cmap else 0,
            "notes":     _cell_to_str(row[cmap["notes"]])     if "notes"     in cmap else "",
        })
    out.sort(key=lambda r: r["date"])
    return out


def _parse_errors(rows: list[list]) -> list[dict]:
    if not rows:
        return []
    header, *body = rows
    cmap = _build_column_map(header, ERROR_FIELD_ALIASES)
    out = []
    for row in body:
        if row is None or all(c is None or c == "" for c in row):
            continue
        no_val = _cell_to_int(row[cmap["no"]]) if "no" in cmap else 0
        date_val = _cell_to_str(row[cmap["date"]]) if "date" in cmap else ""
        if not no_val and not date_val:
            continue
        out.append({
            "no":     no_val,
            "date":   date_val,
            "time":   _cell_to_str(row[cmap["time"]])   if "time"   in cmap else "",
            "cycle":  _cell_to_int(row[cmap["cycle"]])  if "cycle"  in cmap else 0,
            "code":   _cell_to_str(row[cmap["code"]])   if "code"   in cmap else "",
            "type":   _cell_to_str(row[cmap["type"]])   if "type"   in cmap else "",
            "detail": _cell_to_str(row[cmap["detail"]]) if "detail" in cmap else "",
            "cause":  _cell_to_str(row[cmap["cause"]])  if "cause"  in cmap else "",
            "action": _cell_to_str(row[cmap["action"]]) if "action" in cmap else "",
            "result": _cell_to_str(row[cmap["result"]]) if "result" in cmap else "",
            "owner":  _cell_to_str(row[cmap["owner"]])  if "owner"  in cmap else "",
        })
    out.sort(key=lambda r: r.get("no", 0))
    return out


def _pick_latest_xlsx() -> Path:
    xlsxs = sorted(
        [p for p in RAW_DIR.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not xlsxs:
        raise SystemExit(f"data/raw/ 폴더에 .xlsx 파일이 없습니다. ({RAW_DIR})")
    return xlsxs[0]


# ── 시트 로딩: openpyxl 우선, 실패하면 xlwings ────────────────────
def _load_via_openpyxl(src: Path) -> tuple[list[list], list[list], list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(src, data_only=True)
    names = wb.sheetnames
    daily_name  = _find_name(names, DAILY_SHEET_KEYWORDS)
    errors_name = _find_name(names, ERRORS_SHEET_KEYWORDS)

    if daily_name is None:
        raise SystemExit(f"'일일평가' 시트를 찾지 못했습니다. 시트 목록: {names}")

    daily_rows  = [list(r) for r in wb[daily_name].iter_rows(values_only=True)]
    errors_rows = (
        [list(r) for r in wb[errors_name].iter_rows(values_only=True)]
        if errors_name else []
    )
    return daily_rows, errors_rows, names


def _xlwings_sheet_rows(sheet) -> list[list]:
    rng = sheet.used_range
    val = rng.value
    if val is None:
        return []
    if not isinstance(val, list):
        return [[val]]
    if val and not isinstance(val[0], list):
        # 1D 결과 — 단일 행 or 단일 열
        if rng.rows.count == 1:
            return [val]
        return [[v] for v in val]
    return val


def _load_via_xlwings(src: Path) -> tuple[list[list], list[list], list[str]]:
    try:
        import xlwings as xw
    except ImportError as e:
        raise SystemExit(
            "xlwings 미설치. DRM 보호 파일을 읽으려면 'pip install xlwings' 후 재시도. "
            "(Windows + Excel 설치 필수)"
        ) from e

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    try:
        wb = app.books.open(str(src), update_links=False, read_only=True)
        try:
            names = [s.name for s in wb.sheets]
            daily_name  = _find_name(names, DAILY_SHEET_KEYWORDS)
            errors_name = _find_name(names, ERRORS_SHEET_KEYWORDS)

            if daily_name is None:
                raise SystemExit(f"'일일평가' 시트를 찾지 못했습니다. 시트 목록: {names}")

            daily_rows  = _xlwings_sheet_rows(wb.sheets[daily_name])
            errors_rows = (
                _xlwings_sheet_rows(wb.sheets[errors_name])
                if errors_name else []
            )
            return daily_rows, errors_rows, names
        finally:
            wb.close()
    finally:
        app.quit()


def _load_workbook_rows(src: Path) -> tuple[list[list], list[list]]:
    try:
        daily_rows, errors_rows, _ = _load_via_openpyxl(src)
        return daily_rows, errors_rows
    except zipfile.BadZipFile:
        # DRM 래핑 추정 — openpyxl은 zip 구조가 아니라고 거부함
        print("[build] openpyxl 실패 (DRM 추정) → xlwings로 Excel 통한 재시도")
        daily_rows, errors_rows, _ = _load_via_xlwings(src)
        return daily_rows, errors_rows


def main():
    src = _pick_latest_xlsx()
    print(f"[build] 입력 파일: {src.name}")

    daily_rows, errors_rows = _load_workbook_rows(src)
    daily  = _parse_daily(daily_rows)
    errors = _parse_errors(errors_rows)

    out = {
        "generatedAt": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source":      src.name,
        "daily":       daily,
        "errors":      errors,
    }
    OUT_PATH.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[build] 출력: {OUT_PATH.relative_to(ROOT)}  (daily {len(daily)}건, errors {len(errors)}건)")


if __name__ == "__main__":
    main()

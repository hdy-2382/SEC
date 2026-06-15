"""
generate_vendor_template.py
---------------------------
업체에게 보내줄 양식 엑셀(.xlsx)을 자동 생성한다.

생성 위치: data/vendor_template.xlsx
시트 구성:
  - 안내      : 사용 방법, 컬럼 설명
  - 일일평가  : 매일 입력하는 평가 데이터 (헤더 + 예시 2행)
  - 에러로그  : 에러 발생 시마다 1행 추가

build_dashboard_json.py 가 인식하는 컬럼명·시트명을 그대로 사용한다.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "vendor_template.xlsx"


# ── 스타일 ───────────────────────────────────────────────
NAVY = "1A2942"
NAVY_DEEP = "0C1B36"
MUSTARD = "B88A2B"
LINE = "D6D2C4"
BG_ALT = "FAF8F2"
RUST = "8B2E1F"

HEADER_FILL = PatternFill("solid", fgColor=NAVY_DEEP)
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
EXAMPLE_FILL = PatternFill("solid", fgColor=BG_ALT)
EXAMPLE_FONT = Font(name="맑은 고딕", size=10, italic=True, color="7B8087")
BODY_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color=NAVY_DEEP)
SECTION_FONT = Font(name="맑은 고딕", size=12, bold=True, color=NAVY)
NOTE_FONT = Font(name="맑은 고딕", size=10, color="3D4147")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws: Worksheet, row: int, cols: list[str], widths: list[int]):
    for i, (name, width) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=row, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 32


def _style_example_row(ws: Worksheet, row: int, values: list, n_cols: int):
    for i in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=i, value=values[i - 1] if i - 1 < len(values) else "")
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.alignment = LEFT
        cell.border = BORDER


def build_guide_sheet(wb: Workbook):
    ws = wb.create_sheet("안내", 0)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80

    ws["A1"] = "로봇 자동화 양산평가 — 일일 입력 양식"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 30

    rows = [
        ("", ""),
        ("작성 방법", ""),
        ("", "1) 본 파일은 양식이며, 매일 평가가 끝나면 [일일평가] 시트에 1행씩 추가합니다."),
        ("", "2) 에러가 발생한 날에는 [에러로그] 시트에도 해당 에러 1건을 별도로 추가합니다."),
        ("", "3) 파일명은 자유 (예: 양산평가_2026-06-17.xlsx). 한 파일에 누적 기록을 유지하세요."),
        ("", "4) 작성 완료한 파일을 PM에게 전달하면 GitHub의 data/raw/ 폴더에 업로드됩니다."),
        ("", ""),
        ("주의사항", ""),
        ("", "• 시트명(일일평가 / 에러로그)은 변경하지 마세요."),
        ("", "• 헤더 행(파란색)은 그대로 유지하세요. 컬럼 순서·이름이 바뀌면 자동 변환이 실패합니다."),
        ("", "• 날짜는 YYYY-MM-DD 형식 (예: 2026-06-17). 시각은 HH:MM (예: 14:32)."),
        ("", "• 회색 이탤릭 예시 행은 참고용이며, 실제 데이터 입력 시 덮어쓰거나 위에 추가하세요."),
        ("", ""),
        ("핵심 메트릭", ""),
        ("", "• 일일평가 : 그 날 수행한 총 사이클 횟수 (성공+실패 모두 포함)"),
        ("", "• 일일에러 : 그 날 발생한 에러(인시던트) 횟수"),
        ("", "• 연속성공 : 마지막 에러 이후 누적 성공 사이클 수"),
        ("", ""),
        ("에러 상세자료 (선택)", ""),
        ("", "• 상세설명 : 에러로그 행마다 추가로 길게 적고 싶은 분석/경위. 대시보드의 [＋상세] 버튼에서만 보입니다."),
        ("", "• 사진(파일명) : 첨부할 이미지 파일명을 쉼표로 구분해 적습니다. 예) ERR-001_1.jpg, ERR-001_2.jpg"),
        ("", "    └ 사진 파일 자체는 엑셀에 붙이지 말고, 파일명만 적은 뒤 이미지 파일들을 엑셀과 함께 PM에게 전달하세요."),
        ("", "    └ PM이 data/errors/ 폴더에 같은 이름으로 넣으면 [＋상세]에서 표시됩니다. (미입력 시 버튼은 표시되지 않음)"),
        ("", ""),
        ("문의", ""),
        ("", "양식 변경·문제 발생 시 PM에게 문의."),
    ]
    for offset, (lbl, body) in enumerate(rows, start=2):
        a = ws.cell(row=offset, column=1, value=lbl)
        b = ws.cell(row=offset, column=2, value=body)
        if lbl in ("작성 방법", "주의사항", "핵심 메트릭", "문의"):
            a.font = SECTION_FONT
        else:
            a.font = BODY_FONT
        b.font = NOTE_FONT
        b.alignment = LEFT


def build_daily_sheet(wb: Workbook):
    ws = wb.create_sheet("일일평가")
    cols    = ["평가일",     "입실인원",     "주평가내용",                 "일일평가", "일일에러", "연속성공", "비고"]
    widths  = [14,           20,             46,                            12,         12,         12,         24]
    _style_header(ws, 1, cols, widths)
    ws.freeze_panes = "A2"

    examples = [
        ["2026-06-01", "홍길동, 김철수", "JOB 생성 - 픽업 - 적재 사이클 셋업", 42, 0, 42,  "초기 셋업, 안정"],
        ["2026-06-02", "홍길동, 김철수", "사이클 반복 안정성 검증",            78, 0, 120, "정상"],
    ]
    for i, ex in enumerate(examples, start=2):
        _style_example_row(ws, i, ex, len(cols))
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22


def build_errors_sheet(wb: Workbook):
    ws = wb.create_sheet("에러로그")
    cols   = ["No", "발생일",     "시각",  "회차", "코드",     "유형",         "상세",                                "원인",                            "조치",                                 "결과",       "삼성 담당자", "업체 담당자", "상세설명",                              "사진(파일명)"]
    widths = [6,    14,           10,      10,     12,         18,             46,                                    36,                                  44,                                      14,           14,            14,            54,                                      26]
    _style_header(ws, 1, cols, widths)
    ws.freeze_panes = "A2"

    examples = [
        [1, "2026-06-08", "14:32", 358, "ERR-001", "비전 인식 오류",   "픽업 대상 부품의 비전 좌표 인식 실패, 로봇 정지",
         "조도 변화로 카메라 노출값 부적합 추정", "조명 LUX 재조정 + 비전 알고리즘 threshold 보정", "정상복귀", "양희두", "박영희",
         "현장 조도 320→210 LUX로 급감한 구간에서 반복 발생. 노출 보정 후 재현 안 됨. (상세 분석 리포트 별첨)", "ERR-001_1.jpg, ERR-001_2.jpg"],
        [2, "2026-06-17", "11:08", 953, "ERR-002", "그리퍼 그립 실패", "부품 표면 마찰계수 편차로 그리핑 실패, 자동 정지",
         "부품 표면 코팅 편차 추정",            "그리퍼 압력 +5% 조정, 부품 표면 사전 검사 단계 추가", "정상복귀", "김현일", "홍길동",
         "코팅 로트 편차로 마찰계수 0.4→0.28. 압력 상향으로 해결.", "ERR-002_grip.png"],
    ]
    for i, ex in enumerate(examples, start=2):
        _style_example_row(ws, i, ex, len(cols))
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 38


def main():
    wb = Workbook()
    # 기본 생성된 시트 제거
    default_sheet = wb.active
    wb.remove(default_sheet)

    build_guide_sheet(wb)
    build_daily_sheet(wb)
    build_errors_sheet(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[template] 생성 완료: {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
